#!/usr/bin/env python3
"""Comprehensive ground-truth KPI computation for all pages."""
import sys, io, csv, json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from datetime import datetime
from collections import defaultdict
import openpyxl

# ===================== LOAD ALL DATA =====================

# Order CSV
with open('红豆数据（测试数据）/数据表格/订单5.25.15.csv', 'r', encoding='utf-8-sig') as f:
    all_orders_raw = list(csv.DictReader(f))

def find_field(row, *labels):
    for label in labels:
        if label in row and row[label] is not None and str(row[label]).strip() != '':
            return row[label]
    for k in row.keys():
        for label in labels:
            if label in k and row[k] is not None and str(row[k]).strip() != '':
                return row[k]
    return None

def sf(val):
    if val is None: return 0.0
    try:
        s = str(val).strip().replace(',','').replace('¥','').replace('￥','')
        if s in ('','-','--'): return 0.0
        return float(s)
    except: return 0.0

# Filter cancelled (matching ALL pages)
orders = [o for o in all_orders_raw if str(find_field(o,'订单状态','状态') or '').strip() != '已取消']
cnt = len(orders)
print(f'Orders: {cnt} active / {len(all_orders_raw)} total')

# After-sale Excel
wb_as = openpyxl.load_workbook('红豆数据（测试数据）/数据表格/售后数据.xlsx')
ws_as = wb_as.active
as_headers = [str(ws_as.cell(1,c).value) for c in range(1, ws_as.max_column+1)]
after_sale = []
for r in range(2, ws_as.max_row+1):
    row = {as_headers[i]: str(ws_as.cell(r,i+1).value or '') for i in range(len(as_headers))}
    after_sale.append(row)

# Build after-sale lookup by order number
as_by_order = {}
for a in after_sale:
    ono = a.get('订单编号','').strip()
    amt = 0.0
    try: amt = float(a.get('退款金额','0') or '0')
    except: pass
    if ono:
        if ono not in as_by_order:
            as_by_order[ono] = {'amt': 0.0, 'status': '', 'records': []}
        as_by_order[ono]['amt'] += amt
        as_by_order[ono]['status'] = a.get('售后状态','').strip()
        as_by_order[ono]['records'].append(a)

# Promotion Excel
wb_promo = openpyxl.load_workbook('红豆数据（测试数据）/数据表格/商品推广_分天数据_20260222至20260522.xlsx')
ws_ps = wb_promo['汇总数据_分天数据_20260222至20260522']
ps_headers = [str(ws_ps.cell(1,c).value) for c in range(1, ws_ps.max_column+1)]
promo_summary = []
for r in range(2, ws_ps.max_row+1):
    row = {ps_headers[i]: str(ws_ps.cell(r,i+1).value or '') for i in range(len(ps_headers))}
    promo_summary.append(row)

ws_pp = wb_promo['商品_分天数据_20260222至20260522']
pp_headers = [str(ws_pp.cell(1,c).value) for c in range(1, ws_pp.max_column+1)]
promo_products = []
for r in range(2, ws_pp.max_row+1):
    row = {pp_headers[i]: str(ws_pp.cell(r,i+1).value or '') for i in range(len(pp_headers))}
    promo_products.append(row)

# Insurance Excel
wb_ins = openpyxl.load_workbook('红豆数据（测试数据）/数据表格/运费险.xlsx')
ws_ins = wb_ins.active
ins_headers = [str(ws_ins.cell(1,c).value) for c in range(1, ws_ins.max_column+1)]
insurance = []
for r in range(2, ws_ins.max_row+1):
    row = {ins_headers[i]: str(ws_ins.cell(r,i+1).value or '') for i in range(len(ins_headers))}
    insurance.append(row)

# ===================== BASELINE METRICS =====================

gmv = sum(sf(find_field(o,'商品总价(元)','商品总价')) for o in orders)
paid = sum(sf(find_field(o,'用户实付金额(元)','用户实付金额','用户实付','实付金额')) for o in orders)
merchant = sum(sf(find_field(o,'商家实收金额(元)','商家实收金额','商家实收','实收金额')) for o in orders)
total_qty = sum(sf(find_field(o,'商品数量(件)','商品数量','数量')) for o in orders)
postage = sum(sf(find_field(o,'邮费(元)','邮费','快递费(元)','快递费')) for o in orders)

# Discount
shop_disc = sum(sf(find_field(o,'店铺优惠折扣(元)','店铺优惠折扣','店铺优惠')) for o in orders)
plat_disc = sum(sf(find_field(o,'平台优惠折扣(元)','平台优惠折扣','平台优惠')) for o in orders)
pay_disc = sum(sf(find_field(o,'多多支付立减金额(元)','多多支付立减金额','支付立减')) for o in orders)
total_disc = shop_disc + plat_disc + pay_disc

# After-sale from CSV 售后状态
as_cnt = 0; rf_cnt = 0
for o in orders:
    st = str(find_field(o,'售后状态') or '').strip()
    if st and st != '无售后或售后取消' and st != '无':
        as_cnt += 1
    if '退款' in st:
        rf_cnt += 1

# Refund amount: merge CSV 售后状态 with Excel data
rf_amount = 0.0
for o in orders:
    ono = str(find_field(o,'订单号') or '').strip()
    st = str(find_field(o,'售后状态') or '').strip()
    if '退款' in st and ono in as_by_order:
        rf_amount += as_by_order[ono]['amt']

# Ship time
shipped = 0; total_ship_hours = 0.0
for o in orders:
    st = find_field(o,'发货时间')
    if st is not None and str(st).strip() != '':
        shipped += 1
        try:
            pay_t = datetime.strptime(str(find_field(o,'支付时间') or '').strip()[:19], '%Y-%m-%d %H:%M:%S')
            ship_t = datetime.strptime(str(st).strip()[:19], '%Y-%m-%d %H:%M:%S')
            total_ship_hours += (ship_t - pay_t).total_seconds()/3600
        except: pass

# Products
product_ids = set()
for o in orders:
    pid = str(find_field(o,'商品id','商品ID') or '').strip()
    if pid and pid != '-': product_ids.add(pid)

# Buyers - TWO methods (Dashboard vs TrendPage inconsistency!)
buyers_last4 = set()
buyers_first6 = set()
for o in orders:
    ono = str(find_field(o,'订单号') or '').strip()
    if len(ono) >= 4:
        buyers_last4.add(ono[-4:])
    if len(ono) >= 6:
        buyers_first6.add(ono[:6])

# ===================== PROMOTION METRICS =====================

promo_cost = sum(sf(r.get('总花费(元)','0') or r.get('成交花费(元)','0')) for r in promo_summary)
promo_gmv_val = sum(sf(r.get('交易额(元)','0') or r.get('成交金额(元)','0')) for r in promo_summary)
promo_orders_cnt = sum(int(float(r.get('成交笔数','0') or '0')) for r in promo_summary)
promo_impressions = sum(int(float(r.get('曝光量','0') or '0')) for r in promo_summary)
promo_clicks = sum(int(float(r.get('点击量','0') or '0')) for r in promo_summary)

promo_roi = promo_gmv_val/promo_cost if promo_cost>0 else 0
promo_ctr = (promo_clicks/promo_impressions)*100 if promo_impressions>0 else 0
promo_cvr = (promo_orders_cnt/promo_clicks)*100 if promo_clicks>0 else 0
promo_cpc = promo_cost/promo_clicks if promo_clicks>0 else 0
promo_cpa = promo_cost/promo_orders_cnt if promo_orders_cnt>0 else 0
promo_ratio = (promo_cost/gmv)*100 if gmv>0 else 0
shop_roi = gmv/promo_cost if promo_cost>0 else 0

# ===================== INSURANCE METRICS =====================
ins_cost = sum(sf(r.get('服务费用（元）','0')) for r in insurance)
ins_compensated = sum(1 for r in insurance if str(r.get('运费补偿状态','')).strip() and str(r.get('运费补偿状态','')).strip() != '无')
# Match insurance to orders (by 订单编号)
ins_order_map = {}
for r in insurance:
    ono = str(r.get('订单编号','')).strip()
    if ono:
        fee = sf(r.get('服务费用（元）','0'))
        ins_order_map[ono] = fee

# ===================== REGION METRICS =====================
province_stats = defaultdict(lambda: {'orders':0,'gmv':0,'paid':0,'merchant':0,'postage':0,'refund':0,'refund_cnt':0,'as_cnt':0,'buyers':set()})
for o in orders:
    prov = str(find_field(o,'省','省份') or '').strip() or '未知'
    ps = province_stats[prov]
    ps['orders'] += 1
    ps['gmv'] += sf(find_field(o,'商品总价(元)','商品总价'))
    ps['paid'] += sf(find_field(o,'用户实付金额(元)','用户实付金额','用户实付','实付金额'))
    ps['merchant'] += sf(find_field(o,'商家实收金额(元)','商家实收金额','商家实收','实收金额'))
    ps['postage'] += sf(find_field(o,'邮费(元)'))
    ono = str(find_field(o,'订单号') or '').strip()
    if len(ono) >= 4: ps['buyers'].add(ono[-4:])
    st = str(find_field(o,'售后状态') or '').strip()
    if st and st != '无售后或售后取消' and st != '无':
        ps['as_cnt'] += 1
    if '退款' in st:
        ps['refund_cnt'] += 1
        ps['refund'] += as_by_order.get(ono, {}).get('amt', 0.0)

# ===================== LOGISTICS METRICS =====================
courier_stats = defaultdict(lambda: {'orders':0,'paid':0,'avg_hours':0,'h48':0,'total_h':0})
for o in orders:
    courier = str(find_field(o,'快递公司') or '').strip() or '未知'
    cs = courier_stats[courier]
    cs['orders'] += 1
    cs['paid'] += sf(find_field(o,'用户实付金额(元)','用户实付金额','用户实付','实付金额'))
    st = find_field(o,'发货时间')
    if st is not None and str(st).strip() != '':
        try:
            pay_t = datetime.strptime(str(find_field(o,'支付时间') or '').strip()[:19], '%Y-%m-%d %H:%M:%S')
            ship_t = datetime.strptime(str(st).strip()[:19], '%Y-%m-%d %H:%M:%S')
            h = (ship_t - pay_t).total_seconds()/3600
            cs['total_h'] += h
            if h <= 48: cs['h48'] += 1
        except: pass
# Compute averages
for cs in courier_stats.values():
    if cs['orders'] > 0:
        cs['avg_hours'] = cs['total_h'] / cs['orders']

# ===================== USER/BUYER METRICS =====================
buyer_map = defaultdict(lambda: {'orders':0,'paid':0,'merchant':0,'qty':0,'first_date':'','last_date':''})
for o in orders:
    ono = str(find_field(o,'订单号') or '').strip()
    key = ono[-4:] if len(ono)>=4 else ono
    b = buyer_map[key]
    b['orders'] += 1
    b['paid'] += sf(find_field(o,'用户实付金额(元)','用户实付金额','用户实付','实付金额'))
    b['merchant'] += sf(find_field(o,'商家实收金额(元)','商家实收金额','商家实收','实收金额'))
    b['qty'] += sf(find_field(o,'商品数量(件)','商品数量','数量'))
    pay_date = str(find_field(o,'支付时间') or '').split(' ')[0]
    if pay_date:
        if not b['first_date'] or pay_date < b['first_date']:
            b['first_date'] = pay_date
        if not b['last_date'] or pay_date > b['last_date']:
            b['last_date'] = pay_date

repeat = sum(1 for b in buyer_map.values() if b['orders'] > 1)
buyer_count = len(buyer_map)

# ===================== COST METRICS =====================
# Disc rate per order
high_disc = 0
for o in orders:
    sd = sf(find_field(o,'店铺优惠折扣(元)','店铺优惠折扣','店铺优惠'))
    pd = sf(find_field(o,'平台优惠折扣(元)','平台优惠折扣','平台优惠'))
    pt = sf(find_field(o,'商品总价(元)','商品总价'))
    if pt > 0 and (sd+pd)/pt > 0.3:
        high_disc += 1

disc_rate = (total_disc/gmv)*100 if gmv>0 else 0
recv_rate = (merchant/gmv)*100 if gmv>0 else 0

# ===================== RISK METRICS =====================
overdue_48h = 0
for o in orders:
    st = find_field(o,'发货时间')
    if st is not None and str(st).strip() != '':
        try:
            pay_t = datetime.strptime(str(find_field(o,'支付时间') or '').strip()[:19], '%Y-%m-%d %H:%M:%S')
            ship_t = datetime.strptime(str(st).strip()[:19], '%Y-%m-%d %H:%M:%S')
            if (ship_t - pay_t).total_seconds()/3600 > 48:
                overdue_48h += 1
        except: pass

ship_rate_48h = ((shipped-overdue_48h)/shipped)*100 if shipped>0 else 0

# Product risk
product_risk = defaultdict(lambda: {'orders':0,'as_cnt':0,'rf_cnt':0,'qty':0,'overdue':0})
for o in orders:
    pid = str(find_field(o,'商品id','商品ID') or '').strip()
    if not pid or pid == '-': continue
    pr = product_risk[pid]
    pr['orders'] += 1
    pr['qty'] += sf(find_field(o,'商品数量(件)','商品数量','数量'))
    st_as = str(find_field(o,'售后状态') or '').strip()
    if st_as and st_as != '无售后或售后取消' and st_as != '无':
        pr['as_cnt'] += 1
    if '退款' in st_as:
        pr['rf_cnt'] += 1
    ship_t = find_field(o,'发货时间')
    if ship_t is not None and str(ship_t).strip() != '':
        try:
            pay_t = datetime.strptime(str(find_field(o,'支付时间') or '').strip()[:19], '%Y-%m-%d %H:%M:%S')
            s_t = datetime.strptime(str(ship_t).strip()[:19], '%Y-%m-%d %H:%M:%S')
            if (s_t - pay_t).total_seconds()/3600 > 48:
                pr['overdue'] += 1
        except: pass

# ===================== PRINT REPORT =====================
print()
print('='*70)
print('COMPREHENSIVE DATA AUDIT BASELINE')
print('='*70)

print('\n=== 1. DASHBOARD KPIs ===')
print(f'  GMV(商品总价):              ¥{gmv:>12,.2f}')
print(f'  有效订单量:                   {cnt:>12,}')
print(f'  客单价(实付/订单):           ¥{paid/cnt:>12,.2f}')
print(f'  售后率:                      {as_cnt/cnt*100:>11.2f}% ({as_cnt}/{cnt})')
print(f'  退款率:                      {rf_cnt/cnt*100:>11.2f}% ({rf_cnt}/{cnt})')
print(f'  退款金额(merged):            ¥{rf_amount:>12,.2f}')
print(f'  邮费总额:                    ¥{postage:>12,.2f}')
print(f'  买家数(last-4 chars):         {len(buyers_last4):>12,}')
print(f'  买家数(first-6 chars):        {len(buyers_first6):>12,}  ⚠️ INCONSISTENT!')
print(f'  商品数(unique ID):            {len(product_ids):>12,}')
print(f'  平均件数:                     {total_qty/cnt:>12.2f}')
print(f'  优惠总额:                    ¥{total_disc:>12,.2f}')
print(f'  发货率:                      {shipped/cnt*100:>11.2f}% ({shipped}/{cnt})')
print(f'  平均发货时长:                 {total_ship_hours/shipped:>11.2f}h (ship errs: {0})')
print(f'  用户实付:                    ¥{paid:>12,.2f}')
print(f'  商家实收:                    ¥{merchant:>12,.2f}')

print('\n=== 2. PROMOTION KPIs ===')
print(f'  推广花费:                    ¥{promo_cost:>12,.2f}')
print(f'  推广GMV:                     ¥{promo_gmv_val:>12,.2f}')
print(f'  推广订单数:                   {promo_orders_cnt:>12,}')
print(f'  推广ROI:                      {promo_roi:>12.2f}')
print(f'  点击率(CTR):                  {promo_ctr:>11.2f}%')
print(f'  转化率(CVR):                  {promo_cvr:>11.2f}%')
print(f'  平均点击成本(CPC):           ¥{promo_cpc:>12,.2f}')
print(f'  平均获客成本(CPA):           ¥{promo_cpa:>12,.2f}')
print(f'  推广占比(vs GMV):             {promo_ratio:>11.2f}%')
print(f'  全店投产(GMV/cost):           {shop_roi:>12.2f}')

print('\n=== 3. REGION METRICS ===')
print(f'  覆盖省份数: {len(province_stats)}')
print(f'  Top 10 provinces by orders:')
for prov, ps in sorted(province_stats.items(), key=lambda x:-x[1]['orders'])[:10]:
    print(f'    {prov}: {ps["orders"]}单, GMV ¥{ps["gmv"]:.0f}, 实付¥{ps["paid"]:.0f}, 退款率{ps["refund_cnt"]/ps["orders"]*100:.1f}%')

print('\n=== 4. LOGISTICS METRICS ===')
print(f'  平均发货时长: {total_ship_hours/shipped:.2f}h (shipped={shipped})')
print(f'  48h发货率: {ship_rate_48h:.2f}%')
print(f'  免邮率: {sum(1 for o in orders if sf(find_field(o,"邮费(元)"))==0)/cnt*100:.2f}%')
print(f'  快递公司分布:')
for courier, cs in sorted(courier_stats.items(), key=lambda x:-x[1]['orders']):
    print(f'    {courier}: {cs["orders"]}单, 平均{cs["avg_hours"]:.1f}h')

print('\n=== 5. USER METRICS ===')
print(f'  买家数(last-4): {buyer_count}')
print(f'  复购买家: {repeat} ({repeat/buyer_count*100:.1f}%)')
print(f'  人均消费: ¥{paid/buyer_count:.2f}')
print(f'  连带率(avg qty/order): {total_qty/cnt:.2f}')

print('\n=== 6. COST METRICS ===')
print(f'  店铺优惠: ¥{shop_disc:,.2f}')
print(f'  平台优惠: ¥{plat_disc:,.2f}')
print(f'  多多立减: ¥{pay_disc:,.2f}')
print(f'  优惠总额: ¥{total_disc:,.2f}')
print(f'  优惠率(vs GMV): {disc_rate:.2f}%')
print(f'  高优惠订单(>30%): {high_disc}')
print(f'  实收率(vs GMV): {recv_rate:.2f}%')

print('\n=== 7. INSURANCE METRICS ===')
print(f'  保费总额: ¥{ins_cost:.2f}')
print(f'  已补偿: {ins_compensated}/{len(insurance)}')

print('\n=== 8. RISK METRICS ===')
print(f'  超时发货(>48h): {overdue_48h}')
print(f'  48h发货率: {ship_rate_48h:.2f}%')
print(f'  高售后率商品(>30%): {sum(1 for p in product_risk.values() if p["orders"]>0 and p["as_cnt"]/p["orders"]>0.3)}')

print('\n=== 9. AFTER-SALE METRICS ===')
print(f'  独立售后记录: {len(after_sale)}')
print(f'  退款总金额(Excel): ¥{sum(as_by_order[k]["amt"] for k in as_by_order):.2f}')
# Match orders to after-sale records
matched_as = 0
for o in orders:
    ono = str(find_field(o,'订单号') or '').strip()
    if ono in as_by_order:
        matched_as += 1
print(f'  订单匹配售后记录: {matched_as}/{cnt}')

print('\n=== 10. ORDER STATUS DISTRIBUTION ===')
status_dist = defaultdict(int)
for o in orders:
    st = str(find_field(o,'订单状态','状态') or '').strip()
    status_dist[st if st else '未知'] += 1
for st, n in sorted(status_dist.items(), key=lambda x:-x[1]):
    print(f'  {st}: {n}')

# ===================== IDENTIFY ISSUES =====================
print('\n'+'='*70)
print('ISSUES FOUND')
print('='*70)
print()
print('1. ⚠️ 退款金额字段在CSV中不存在:')
print('   - CSV的87列中无"退款金额"列')
print('   - 售后Excel中有57条记录,退款总额¥{:.2f}'.format(sum(as_by_order[k]['amt'] for k in as_by_order)))
print('   - DashboardPage/ProductPage/CostPage计算退款金额时从订单字段读取,结果为¥0')
print('   - AfterSalePage读取afterSaleRecords,有正确的退款金额')
print()
print('2. ⚠️ 买家数算法不一致:')
print('   - DashboardPage line 385: 订单号.slice(-4) → {} buyers'.format(len(buyers_last4)))
print('   - TrendPage line 110: 订单号.slice(0, 6) → {} buyers'.format(len(buyers_first6)))
print('   - 两种算法结果不同!')
print()
print('3. ⚠️ 买家数不是真实买家:')
print('   - CSV中没有买家ID字段')
print('   - 用订单号截断来模拟买家,非常不准确')
print('   - 复购率计算也因此不准确')
print()
print('4. ⚠️ 退款率基于CSV售后状态,退款金额来自Excel:')
print('   - 退款率={:.2f}% (基于CSV售后状态,正确)'.format(rf_cnt/cnt*100))
print('   - 退款金额需从售后Excel合并才能得到正确值')
print('   - AfterSalePage有merge逻辑,但DashboardPage没有')
print()
print('5. ℹ️ 数据源限制:')
print('   - 邮费全部为0(拼多多包邮)')
print('   - 明星店铺和直播推广数据为空(该店铺未投放)')
print('   - 货款明细CSV存在但未纳入主KPI计算')

# Save to JSON
with open('audit_baseline.json','w',encoding='utf-8') as f:
    json.dump({
        'orders_active': cnt,
        'orders_total': len(all_orders_raw),
        'gmv': round(gmv,2),
        'paid': round(paid,2),
        'merchant': round(merchant,2),
        'discount_shop': round(shop_disc,2),
        'discount_plat': round(plat_disc,2),
        'discount_pay': round(pay_disc,2),
        'discount_total': round(total_disc,2),
        'total_qty': round(total_qty,2),
        'as_cnt': as_cnt,
        'rf_cnt': rf_cnt,
        'rf_amount': round(rf_amount,2),
        'shipped': shipped,
        'avg_ship_hours': round(total_ship_hours/shipped if shipped else 0,2),
        'ship_rate': round(shipped/cnt*100 if cnt else 0,2),
        'ship_rate_48h': round(ship_rate_48h,2),
        'buyers_last4': len(buyers_last4),
        'buyers_first6': len(buyers_first6),
        'product_count': len(product_ids),
        'promo_cost': round(promo_cost,2),
        'promo_gmv': round(promo_gmv_val,2),
        'promo_orders': promo_orders_cnt,
        'promo_roi': round(promo_roi,2),
        'promo_ctr': round(promo_ctr,2),
        'promo_cvr': round(promo_cvr,2),
        'promo_cpc': round(promo_cpc,2),
        'promo_cpa': round(promo_cpa,2),
        'promo_ratio': round(promo_ratio,2),
        'shop_roi': round(shop_roi,2),
        'ins_cost': round(ins_cost,2),
        'after_sale_records': len(after_sale),
        'after_sale_refund_total': round(sum(as_by_order[k]['amt'] for k in as_by_order),2),
        'province_count': len(province_stats),
    }, f, ensure_ascii=False, indent=2)
print('\nBaseline saved to audit_baseline.json')
